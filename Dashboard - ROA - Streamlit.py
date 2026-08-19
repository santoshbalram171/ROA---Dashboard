import streamlit as st
import pandas as pd
import os
import shutil
import time
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font

st.set_page_config(page_title="ROA Dashboard", layout="wide")

# --- BEVEILIGING CONFIGURATIE ---
GEBRUIKERS_LIJST = {
    'A': 's',
    'eqh_user': 'Welkom01'
}

# --- CONFIGURATIE & BASISINSTELLINGEN ---
START_VOORRAAD_BASIS = {'UTR': 52, 'T11': 26, 'T12': 52, '2T12': 52, 'TW': 52, 'TWA': 52}
CUSTOM_START_VOORRAAD = {'UTR': 52, 'T11': 26, 'T12': 52, '2T12': 52, 'TW': 52, 'TWA': 52}
alle_excel_kolommen = ['UTR', 'T11', 'T12', '2T12', 'TW', 'TWA']
balk_kolommen = ['UTR', 'T11', 'T12', '2T12', 'TW', 'TWA']

GEWENST_PAD = r"C:\Users\santo\Desktop\Python Script\Trolley Pictures"
AFBEELDING_CONFIG = {
    'UTR': 'UTR.jpg', 'T12': 'T12.jpg', 'TW': 'TW.jpg',
    'T11': 'T11.jpg', '2T12': '2T12.jpg', 'TWA': 'TWA.jpg'
}

if not os.path.exists("assets"):
    os.makedirs("assets")

for type_naam, bestandsnaam in AFBEELDING_CONFIG.items():
    bron_pad = os.path.join(GEWENST_PAD, bestandsnaam)
    doel_pad = os.path.join("assets", bestandsnaam)
    if os.path.exists(bron_pad):
        try:
            shutil.copy(bron_pad, doel_pad)
        except Exception:
            pass

BLOK_TIJDEN = {
    '1': datetime.time(6, 0, 0), '2': datetime.time(7, 25, 0), '3': datetime.time(8, 50, 0),
    '4': datetime.time(10, 15, 0), '5': datetime.time(11, 40, 0), '6': datetime.time(13, 5, 0),
    '7': datetime.time(14, 30, 0), '8': datetime.time(15, 55, 0), '9': datetime.time(17, 20, 0),
    '10': datetime.time(18, 45, 0), '11': datetime.time(20, 10, 0), '12': datetime.time(21, 35, 0)
}

# --- SESSION STATE INITIALISATIE ---
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False

if 'tabel_data' not in st.session_state:
    df_init = pd.DataFrame(columns=['Block', 'Vluchttijd', 'Vluchtnummer'] + alle_excel_kolommen + ['Total'])
    st.session_state.tabel_data = df_init.to_dict('records')

if 'handmatige_min' not in st.session_state:
    berekende_handmatige_min = {}
    for col in balk_kolommen:
        custom_val = CUSTOM_START_VOORRAAD.get(col)
        if custom_val != '' and custom_val is not None:
            try:
                verschil = START_VOORRAAD_BASIS[col] - int(custom_val)
                berekende_handmatige_min[col] = max(0, verschil)
            except ValueError:
                berekende_handmatige_min[col] = 0
        else:
            berekende_handmatige_min[col] = 0
    st.session_state.handmatige_min = berekende_handmatige_min

if 'handmatige_plus' not in st.session_state:
    st.session_state.handmatige_plus = {col: 0 for col in balk_kolommen}

if 'absolute_voorraad' not in st.session_state:
    st.session_state.absolute_voorraad = START_VOORRAAD_BASIS

if 'actieve_timers' not in st.session_state:
    st.session_state.actieve_timers = []

if 'actie_historie' not in st.session_state:
    st.session_state.actie_historie = []

if 'handmatige_vrijgaven' not in st.session_state:
    st.session_state.handmatige_vrijgaven = []

if 'blok_klik_teller' not in st.session_state:
    st.session_state.blok_klik_teller = {}

# --- FUNCTIES ---
def probeer_getal(waarde):
    try:
        return float(waarde) if '.' in str(waarde) else int(waarde)
    except ValueError:
        return waarde

def is_blok_vrijgegeven(df, blok_naam, handmatige_vrijgaven=None):
    if df.empty:
        return True
    blok_str = str(blok_naam).replace("🕒", "").replace("Block", "").strip()
    if handmatige_vrijgaven and blok_str in handmatige_vrijgaven:
        return True
    if blok_str == '1':
        return True
    nu_tijd = datetime.datetime.now().time()
    if blok_str in BLOK_TIJDEN and nu_tijd >= BLOK_TIJDEN[blok_str]:
        return True
    try:
        blok_nr = int(blok_str)
        vorig_blok_str = str(blok_nr - 1)
        totaal_vorig_blok = df[df['Block'].astype(str).str.replace("🕒", "").str.replace("Block", "").str.strip() == vorig_blok_str]['Total'].sum()
        if totaal_vorig_blok == 0:
            return True
    except ValueError:
        pass
    return False

def bepaal_voorraad_status_en_kleur(actuele_waarde, basis_waarde):
    percentage = (actuele_waarde / basis_waarde) * 100 if basis_waarde != 0 else 100
    if percentage >= 60:
        return "Groen", "Normaal"
    elif percentage >= 40:
        return "Geel", "Let op: Lage voorraad"
    else:
        return "Rood", "WAARSCHUWING: Onvoldoende voorraad!"

ACTIEF_LOG_BESTAND = "trolley_logging.xlsx"

def log_klik_naar_excel(trolley_type, block, vluchtnummer, tijd_nu, actuele_voorraad=None, basis_voorraad=None, status="In gebruik", huidige_plus_stand=0):
    nieuwe_datum = tijd_nu.strftime('%Y-%m-%d')
    nieuwe_tijd = tijd_nu.strftime('%H:%M:%S')
    
    if trolley_type in ['MUTR-CCL', 'UTR', 'T12', 'TW', 'OIS', 'A-OIS']:
        equipment = 'EUR'
    elif trolley_type in ['T11', '2T12', 'TWA']:
        equipment = 'KLC'
    else:
        equipment = 'ONBEKEND'

    v_status, waarschuwing = "Groen", "Geen"
    if actuele_voorraad is not None and basis_voorraad is not None:
        v_status, waarschuwing = bepaal_voorraad_status_en_kleur(actuele_voorraad, basis_voorraad)

    log_doel_voorraad = actuele_voorraad + huidige_plus_stand if actuele_voorraad is not None else 'N.v.t.'

    nieuw_record = {
        'Datum': [nieuwe_datum], 'Tijd': [nieuwe_tijd], 'Trolley Type': [trolley_type],
        'Equipment': [equipment], 'Block': [block], 'Vluchtnummer': [vluchtnummer],
        'Status': [status], 'Voorraad Niveau': [actuele_voorraad if actuele_voorraad is not None else 'N.v.t.'],
        'Verwachte Voorraad': [log_doel_voorraad], 'Waarschuwing': [waarschuwing]
    }
    df_nieuw = pd.DataFrame(nieuw_record)
    kolom_volgorde = ['Datum', 'Tijd', 'Trolley Type', 'Equipment', 'Block', 'Vluchtnummer', 'Status', 'Voorraad Niveau', 'Verwachte Voorraad', 'Waarschuwing']
    df_nieuw = df_nieuw[kolom_volgorde]

    try:
        if os.path.exists(ACTIEF_LOG_BESTAND):
            df_bestaand = pd.read_excel(ACTIEF_LOG_BESTAND)
            for col in kolom_volgorde:
                if col not in df_bestaand.columns:
                    df_bestaand[col] = None
            df_totaal = pd.concat([df_bestaand, df_nieuw], ignore_index=True)
            df_totaal = df_totaal[kolom_volgorde]
            df_totaal.to_excel(ACTIEF_LOG_BESTAND, index=False)
        else:
            df_nieuw.to_excel(ACTIEF_LOG_BESTAND, index=False)

        wb = openpyxl.load_workbook(ACTIEF_LOG_BESTAND)
        ws = wb.active
        fill_groen = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        font_groen = Font(color="FFFFFF", bold=True)
        fill_geel = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
        font_geel = Font(color="2C3E50", bold=True)
        fill_rood = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        font_rood = Font(color="FFFFFF", bold=True)

        for row in range(2, ws.max_row + 1):
            w_val = str(ws.cell(row=row, column=10).value)
            if "WAARSCHUWING" in w_val:
                f, fn = fill_rood, font_rood
            elif "Let op" in w_val:
                f, fn = fill_geel, font_geel
            else:
                f, fn = fill_groen, font_groen
            for col_idx in [8, 9]:
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = f
                cell.font = fn
        wb.save(ACTIEF_LOG_BESTAND)
    except Exception as e:
        print(f"Log fout: {e}")

# --- LOGIN SCHERM ---
if not st.session_state.logged_in:
    st.markdown("<h2 style='text-align: center;'>ROA Dashboard Login</h2>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        username = st.text_input("Gebruikersnaam")
        password = st.text_input("Wachtwoord", type="password")
        if st.button("Inloggen", use_container_width=True):
            u_clean = username.strip().lower()
            geboekte = {k.lower(): v for k, v in GEBRUIKERS_LIJST.items()}
            if u_clean in geboekte and password.strip() == geboekte[u_clean]:
                st.session_state.logged_in = True
                st.rerun()
            else:
                st.error("Onjuiste gebruikersnaam of wachtwoord!")
    st.stop()

# --- TIMER CLEANUP ---
nu_time = time.time()
st.session_state.actieve_timers = [t for t in st.session_state.actieve_timers if nu_time < t['vrijgave_tijd']]

# --- HOOFDSCHERM DASHBOARD ---
st.markdown(f"<div style='text-align: right; font-weight: bold;'>{datetime.datetime.now().strftime('%d-%m-%Y | %H:%M:%S')}</div>", unsafe_allow_html=True)
st.markdown("<h1 style='text-align: center;'>Beschikbare Vlucht Informatie - ROA</h1>", unsafe_allow_html=True)

# Voorraad balken tonen
st.markdown("### Actuele voorraad in EQH-afdeling:")
cols_balk = st.columns(6)
balk_data_stats = {}
for idx, col_name in enumerate(balk_kolommen):
    basis = st.session_state.absolute_voorraad.get(col_name, START_VOORRAAD_BASIS[col_name])
    t_plus = st.session_state.handmatige_plus.get(col_name, 0)
    in_wacht = sum(1 for t in st.session_state.actieve_timers if t['type'] == col_name)
    afgeboekt = st.session_state.handmatige_min.get(col_name, 0)
    doel = max(0, basis + t_plus - afgeboekt)
    actueel = max(0, doel - in_wacht)
    balk_data_stats[col_name] = (actueel, doel, basis)

    pct = (actueel / basis * 100) if basis != 0 else 100
    if pct >= 60:
        bg, fg = "#27ae60", "white"
    elif pct >= 40:
        bg, fg = "#f1c40f", "#2c3e50"
    else:
        bg, fg = "#e74c3c", "white"

    with cols_balk[idx]:
        img_file = AFBEELDING_CONFIG.get(col_name, '')
        img_path = os.path.join("assets", img_file)
        st.markdown(f"""
            <div style='background-color: {bg}; color: {fg}; padding: 10px; border-radius: 8px; text-align: center; margin-bottom: 10px;'>
                <b>{col_name}</b><br>
                <span style='font-size: 18px;'>{actueel} / {doel}</span>
            </div>
        """, unsafe_allow_html=True)
        if os.path.exists(img_path):
            st.image(img_path, width=70)

st.markdown("---")

# Besturingselementen: Upload, Filter, Undo
c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
with c1:
    uploaded_file = st.file_uploader("Upload Excel Schema", type=["xlsx", "xls"])
    if uploaded_file is not None:
        try:
            df_up = pd.read_excel(uploaded_file)
            for col in alle_excel_kolommen + ['Total']:
                if col in df_up.columns:
                    df_up[col] = df_up[col].fillna(0).astype(int)
            if 'Block' in df_up.columns:
                df_up['Block'] = df_up['Block'].astype(str).str.replace("🕒", "").str.replace("Block", "").str.strip()
            if 'IsVrijgegeven' not in df_up.columns:
                df_up['IsVrijgegeven'] = False
            st.session_state.tabel_data = df_up.to_dict('records')
            st.session_state.actieve_timers = []
            st.session_state.actie_historie = []
            st.session_state.handmatige_vrijgaven = []
            st.session_state.blok_klik_teller = {}
            st.success("Bestand succesvol ingeladen!")
        except Exception as e:
            st.error(f"Fout bij inlezen: {e}")

df_current = pd.DataFrame(st.session_state.tabel_data)

with c2:
    view_mode = st.selectbox("Kies Visualisatie", ['Vlucht Details (Standaard)', 'Totaal Per Blok'], index=0)
    mode_val = 'details' if view_mode == 'Vlucht Details (Standaard)' else 'totaal_block'

with c3:
    unieke_blokken_lijst = ['Alle', 'Ochtend', 'Middag']
    if not df_current.empty and 'Block' in df_current.columns:
        b_vals = sorted([str(b) for b in df_current['Block'].unique() if pd.notna(b)], key=probeer_getal)
        unieke_blokken_lijst.extend(b_vals)
    selected_block = st.selectbox("Kies een Block / Dienst", unieke_blokken_lijst)

with c4:
    st.write("")
    st.write("")
    if st.button("↩️ Ongedaan Maken"):
        historie = [a for a in st.session_state.actie_historie if (time.time() - a.get('timestamp', 0)) <= 900.0]
        if historie:
            laatste = historie.pop()
            st.session_state.actie_historie = historie
            t_col, t_bl, t_fl = laatste['trolley_type'], laatste['block'], laatste['vluchtnummer']
            
            if t_fl != "Cumulatief blok":
                m_match = df_current[(df_current['Block'] == t_bl) & (df_current['Vluchtnummer'].astype(str).str.strip() == t_fl)]
            else:
                m_match = df_current[df_current['Block'] == t_bl]

            if not m_match.empty:
                m_idx = m_match.index[0]
                df_current.at[m_idx, t_col] = int(df_current.at[m_idx, t_col]) + 1
                df_current.at[m_idx, 'Total'] = df_current.loc[m_idx, alle_excel_kolommen].astype(int).sum()
                if t_col in balk_kolommen:
                    st.session_state.handmatige_plus[t_col] = max(0, st.session_state.handmatige_plus[t_col] - 1)
                    for i in reversed(range(len(st.session_state.actieve_timers))):
                        if st.session_state.actieve_timers[i]['type'] == t_col:
                            st.session_state.actieve_timers.pop(i)
                            break
                st.session_state.tabel_data = df_current.to_dict('records')
                st.success(f"Actie ongedaan gemaakt voor {t_col}!")
                st.rerun()

# Filter dataframe
df_filtered = df_current.copy()
if not df_filtered.empty and 'Block' in df_filtered.columns:
    if selected_block == 'Ochtend':
        df_filtered = df_filtered[df_filtered['Block'].isin(['1','2','3','4','5','6'])]
    elif selected_block == 'Middag':
        df_filtered = df_filtered[df_filtered['Block'].isin(['7','8','9','10','11','12'])]
    elif selected_block and selected_block != 'Alle':
        df_filtered = df_filtered[df_filtered['Block'] == str(selected_block).strip()]

# Bereid data voor weergave
if mode_val == 'totaal_block':
    if not df_filtered.empty:
        agg_d = {col: 'sum' for col in alle_excel_kolommen + ['Total']}
        df_display = df_filtered.groupby('Block', as_index=False).agg(agg_d)
        df_display['sort_key'] = df_display['Block'].apply(probeer_getal)
        df_display = df_display.sort_values(by='sort_key').drop(columns=['sort_key'])
        df_display['Block'] = df_display['Block'].apply(lambda b: f"Block {b} 🕒" if not is_blok_vrijgegeven(df_current, b, st.session_state.handmatige_vrijgaven) else f"Block {b}")
    else:
        df_display = pd.DataFrame(columns=['Block'] + alle_excel_kolommen + ['Total'])
else:
    if not df_filtered.empty:
        df_display = df_filtered.copy()
        df_display['Block'] = df_display['Block'].apply(lambda b: f"Block {b} 🕒" if not is_blok_vrijgegeven(df_current, b, st.session_state.handmatige_vrijgaven) else f"Block {b}")
    else:
        df_display = pd.DataFrame(columns=['Block', 'Vluchttijd', 'Vluchtnummer'] + alle_excel_kolommen + ['Total'])

st.markdown("### Tabel Overzicht")
st.dataframe(df_display, use_container_width=True, hide_index=True)

# Actie interface voor trolleys afboeken
st.markdown("---")
st.markdown("### Trolley Actie Registratie")

col_act1, col_act2, col_act3 = st.columns(3)
with col_act1:
    actief_blokken_lijst = sorted(list(df_current['Block'].unique()), key=probeer_getal) if not df_current.empty else []
    kie_block = st.selectbox("Selecteer Block", actief_blokken_lijst if actief_blokken_lijst else ['1'])

with col_act2:
    if not df_current.empty and kie_block:
        vluchten_in_block = df_current[df_current['Block'].astype(str) == str(kie_block)]['Vluchtnummer'].tolist()
    else:
        vluchten_in_block = []
    kie_vlucht = st.selectbox("Selecteer Vluchtnummer", vluchten_in_block if vluchten_in_block else ['Cumulatief blok'])

with col_act3:
    kie_trolley = st.selectbox("Selecteer Trolley Type", alle_excel_kolommen)

if st.button("Trolley Afboeken / In Gebruik Nemen", type="primary"):
    b_str = str(kie_block).strip()
    if not is_blok_vrijgegeven(df_current, b_str, st.session_state.handmatige_vrijgaven):
        t_data = st.session_state.blok_klik_teller.get(b_str, {'tijd': 0, 'aantal': 0})
        if time.time() - t_data.get('tijd', 0) > 5.0:
            t_data = {'tijd': time.time(), 'aantal': 1}
        else:
            t_data['aantal'] += 1
            t_data['tijd'] = time.time()
        st.session_state.blok_klik_teller[b_str] = t_data

        if t_data['aantal'] >= 3:
            if b_str not in st.session_state.handmatige_vrijgaven:
                st.session_state.handmatige_vrijgaven.append(b_str)
            st.session_state.blok_klik_teller[b_str] = {'tijd': 0, 'aantal': 0}
            st.success(f"Block {b_str} is succesvol vrijgegeven!")
            st.rerun()
        else:
            rest = 3 - t_data['aantal']
            st.warning(f"Block {b_str} is vergrendeld. Klik nog {rest}x op de knop om dit blok vrij te geven.")
    else:
        bron_idx = None
        gebruikte_block = b_str
        
        master_m = df_current[(df_current['Block'].astype(str) == b_str) & (df_current['Vluchtnummer'].astype(str).str.strip() == str(kie_vlucht))]
        if not master_m.empty:
            m_idx = master_m.index[0]
            if int(df_current.at[m_idx, kie_trolley]) > 0:
                bron_idx = m_idx

        if bron_idx is None:
            alle_b_ges = sorted(df_current['Block'].unique().tolist(), key=probeer_getal)
            try:
                h_pos = alle_b_ges.index(b_str)
            except ValueError:
                h_pos = 0
            for nxt_i in range(h_pos + 1, len(alle_b_ges)):
                nxt_b = alle_b_ges[nxt_i]
                if is_blok_vrijgegeven(df_current, nxt_b, st.session_state.handmatige_vrijgaven):
                    lat_m = df_current[(df_current['Block'].astype(str) == str(nxt_b)) & (df_current[kie_trolley] > 0)]
                    if not lat_m.empty:
                        bron_idx = lat_m.index[0]
                        gebruikte_block = str(nxt_b)
                        break

        if bron_idx is not None:
            cur_val = int(df_current.at[bron_idx, kie_trolley])
            if cur_val > 0:
                df_current.at[bron_idx, kie_trolley] = cur_val - 1
                if kie_trolley in balk_kolommen:
                    st.session_state.handmatige_plus[kie_trolley] += 1
                    st.session_state.actieve_timers.append({'type': kie_trolley, 'vrijgave_tijd': time.time() + 900.0})

                st.session_state.actie_historie.append({
                    'trolley_type': kie_trolley,
                    'block': gebruikte_block,
                    'vluchtnummer': str(kie_vlucht),
                    'timestamp': time.time()
                })

                akt_v, doel_v, bas_v = balk_data_stats[kie_trolley]
                log_klik_naar_excel(
                    trolley_type=kie_trolley,
                    block=gebruikte_block,
                    vluchtnummer=str(kie_vlucht),
                    tijd_nu=datetime.datetime.now(),
                    actuele_voorraad=akt_v,
                    basis_voorraad=bas_v,
                    status="In gebruik",
                    huidige_plus_stand=sum(1 for t in st.session_state.actieve_timers if t['type'] == kie_trolley)
                )
                df_current.at[bron_idx, 'Total'] = df_current.loc[bron_idx, alle_excel_kolommen].astype(int).sum()
                st.session_state.tabel_data = df_current.to_dict('records')
                st.success(f"Trolley {kie_trolley} succesvol afgeboekt!")
                st.rerun()
        else:
            st.error("Geen voorraad beschikbaar van dit type in dit of latere vrijgegeven blokken!")
